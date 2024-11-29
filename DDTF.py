
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime

# Path to Excel file
excel_file = "C:\Users\USER\Downloads\test_data.xlsx"

# Load Excel file
workbook = load_workbook(excel_file)
sheet = workbook.active

# Initialize WebDriver
driver = webdriver.Chrome()  # Ensure chromedriver is in PATH
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

# Iterate through Excel rows
for row in range(2, sheet.max_row + 1):  # Start from the second row (row 1 is the header)
    username = sheet.cell(row=row, column=2).value  # Username in column 2
    password = sheet.cell(row=row, column=3).value  # Password in column 3

    try:
        # Locate username and password fields, login button
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "username"))
        ).clear()
        driver.find_element(By.NAME, "username").send_keys(username)

        driver.find_element(By.NAME, "password").clear()
        driver.find_element(By.NAME, "password").send_keys(password)

        driver.find_element(By.XPATH, "//button[@type='submit']").click()

        # Check for successful login (e.g., dashboard is loaded)
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[text()='Dashboard']"))
            )
            result = "Test Passed"
        except:
            result = "Test Failed"
        
        # Logout if login was successful
        if result == "Test Passed":
            driver.find_element(By.XPATH, "//span[text()='Paul Collings']").click()
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//a[text()='Logout']"))
            ).click()
    except Exception as e:
        result = f"Test Failed ({str(e)})"
    
    # Write the result back to Excel
    sheet.cell(row=row, column=7).value = result  # Test Result in column 7
    sheet.cell(row=row, column=4).value = datetime.now().strftime("%d/%m/%Y")  # Date
    sheet.cell(row=row, column=5).value = datetime.now().strftime("%H:%M:%S")  # Time

# Save the updated Excel file
workbook.save(excel_file)

# Close the browser
driver.quit()